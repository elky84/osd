import openpyxl as Excel
import jinja2
import os
import re
import converter
import config
import validator
from resources import resources

def primaryKey(schemaSet):
    id = [x for x in schemaSet if x['type'].startswith('*')]
    if not id:
        return None
    elif len(id) > 1:
        stringify = ', '.join([x['name'] for x in id])
        raise Exception(f"기본키는 2개 이상 존재할 수 없습니다. ({stringify})")
    else:
        return id[0]

def indexKey(schemaSet):
    matches = {x['name']: re.match(r'^\((?P<type>.*)\)$', x['type']) for x in schemaSet}
    id = [x for x, match in matches.items() if match is not None]
    id = [x for x in schemaSet if x['name'] in id]
    if not id:
        return None
    elif len(id) > 1:
        stringify = ', '.join([x['name'] for x in id])
        raise Exception(f"그룹키는 2개 이상 존재할 수 없습니다. ({stringify})")
    else:
        return id[0]

def equalsSchema(schema1, schema2):
    if len(schema1) != len(schema2):
        return False

    for i in range(len(schema1)):
        data = schema1[i]
        for key in data:
            if key not in schema2[i]:
                return False

            diff = set(schema1[i].items()) - set(schema2[i].items())
            if len(diff) > 0:
                return False

    return True

def load(path):
    schemaDict = {}
    dataDict = {}

    workbook = Excel.load_workbook(path, data_only=True)
    for sheet in workbook.worksheets:
        sheetName = sheet.title
        if sheetName.startswith('#'):
            continue

        tableName = sheetName
        if '.' in tableName and not tableName.startswith('.') and not tableName.endswith('.'):
            tableName, _ = sheetName.split('.')

        beginIndex = 1
        while sheet[beginIndex][0].value.startswith('#'):
            beginIndex = beginIndex + 1

        schema = []
        data = []
        commentColumns = []
        for col in sheet[beginIndex]:
            columnName = col.value
            if columnName is None or columnName.startswith('!'):
                break

            if columnName.startswith('#'):
                commentColumns.append(col.column)
                continue

            dtype = sheet[beginIndex+1][col.column-1].value
            usage = sheet[beginIndex+2][col.column-1].value

            schema.append({'name': columnName, 'type': dtype, 'usage': usage})

        
        data = []
        for row in range(sheet.max_row - (beginIndex+2)):
            additionalColumn = 0
            dataSet = {}
            for column, columnName in enumerate([x['name'] for x in schema]):
                while column+additionalColumn+1 in commentColumns:
                    additionalColumn = additionalColumn + 1

                columnName = schema[column]['name']
                dataSet[columnName] = sheet.cell(row+beginIndex+3, column+1+additionalColumn).value
                if type(dataSet[columnName]) is str:
                    dataSet[columnName] = dataSet[columnName].replace('_x000D_', '').replace('\xa0', '')

            if all([x is None for x in dataSet.values()]):
                break
            
            data.append(dataSet)

        if tableName in schemaDict:
            if not equalsSchema(schemaDict[tableName], schema):
                raise Exception(f'{sheetName}의 스키마가 올바르지 않습니다. 합쳐질 테이블과 같은 형식의 스키마여야 합니다.')
        else:
            schemaDict[tableName] = schema

        if tableName in dataDict:
            dataDict[tableName] = dataDict[tableName] + data
        else:
            dataDict[tableName] = data

    return schemaDict, dataDict

def loadEnum(path):
    workbook = Excel.load_workbook(path, data_only=True)
    enumSet = {}
    for sheet in workbook.worksheets:
        if sheet.title.startswith('#'):
            continue
        
        enumSet[sheet.title] = {}

        for row in range(sheet.max_row - 1):
            name = sheet.cell(row+2, 1).value
            if not name:
                continue
            
            desc = sheet.cell(row+2, 2).value
            enumSet[sheet.title][name] = desc

    return enumSet

def loadConst(path):
    workbook = Excel.load_workbook(path, data_only=True)
    constSet = {}
    for sheet in workbook.worksheets:
        if sheet.title.startswith('#'):
            continue

        constSet[sheet.title] = {}

        for row in range(sheet.max_row):
            name = sheet.cell(row+1, 1).value
            if not name:
                continue

            if name.startswith('#'):
                continue

            usage = sheet.cell(row+1, 2).value
            type = sheet.cell(row+1, 3).value
            value = sheet.cell(row+1, 4).value
            constSet[sheet.title][name] = {'usage': usage, 'type': type, 'value': value}

    return constSet

def loads(directory, callback=None):
    schemaDict = {}
    dataDict = {}
    files = [x for x in os.listdir(directory) if x.endswith('.xlsx') and not x.startswith('~')]

    prefixs = [config.file[x]['prefix'] for x in config.file if config.file[x]['prefix'] is not None]
    for prefix in prefixs:
        files = [x for x in files if not x.lower().startswith(prefix)]

    suffixs = [config.file[x]['suffix'] for x in config.file if config.file[x]['suffix'] is not None]
    for suffix in suffixs:
        files = [x for x in files if not x.lower().endswith(suffix)]

    size = len(files)

    progress = 0
    for file in files:
        path = os.path.join(directory, file)

        schema, data = load(path)
        schemaDict.update(schema)
        dataDict.update(data)
        progress = progress + 1

        if callback:
            callback(file, int(progress * 100 / size))

    return schemaDict, dataDict

def loadEnums(directory, callback=None):
    files = [x for x in os.listdir(directory) if not x.startswith('~')]
    if config.file['enum']['prefix'] is not None:
        files = [x for x in files if x.lower().startswith(config.file['enum']['prefix'])]
    
    if config.file['enum']['suffix'] is not None:
        files = [x for x in files if x.lower().endswith(config.file['enum']['suffix'])]

    enumDict = {}
    for file in files:
        path = os.path.join(directory, file)

        for enumName, enumSet in loadEnum(path).items():
            if enumName in enumDict:
                raise Exception(f'{enumName}은 중복 정의되었습니다.')
            enumDict[enumName] = enumSet

    return enumDict

def loadConsts(directory, callback=None):
    files = [x for x in os.listdir(directory) if not x.startswith('~')]
    if config.file['const']['prefix'] is not None:
        files = [x for x in files if x.lower().startswith(config.file['const']['prefix'])]
    
    if config.file['const']['suffix'] is not None:
        files = [x for x in files if x.lower().endswith(config.file['const']['suffix'])]
        
    constDict = {}
    for file in files:
        path = os.path.join(directory, file)

        for constName, constSet in loadConst(path).items():
            if constName in constDict:
                raise Exception(f'{constName}은 중복 정의되었습니다.')
            constDict[constName] = constSet

    return constDict

def relationshipType(type):
    if not validator.isRelation(type):
        return None

    resources.schemaDict = {name: schemaSet for name, schemaSet in resources.schemaDict.items()}
    type = converter.remove(type)
    splitted = type.split('.')
    if len(splitted) == 1:
        if type not in resources.schemaDict:
            raise Exception(f'{type}은 정의되지 않은 테이블입니다.')

        id = primaryKey(resources.schemaDict[type]) or indexKey(resources.schemaDict[type])
        if not id:
            raise Exception(f'{type}은 기본키가 정의되지 않은 타입입니다.')

        id = converter.remove(id['type'], relation=False)
        if validator.isRelation(id):
            return relationshipType(id)
        else:
            return id

    elif len(splitted) == 2:
        namespace, member = splitted
        if namespace not in resources.schemaDict:
            raise Exception(f'{namespace}은 정의되지 않은 테이블입니다.')

        found = [x for x in resources.schemaDict[namespace] if x['name'] == member]
        if not found:
            raise Exception(f"{member}은 {namespace}에 존재하지 않는 컬럼입니다.")

        member = found[0]
        id = converter.remove(member['type'], relation=False)
        if validator.isRelation(id):
            return relationshipType(id)
        else:
            return id
    else:
        raise Exception(f'{type}는 올바른 관계형식이 아닙니다.')

# (...)
def groupType(type):
    match = re.match(config.regex['group'], type)
    if match:
        return match.groupdict()['type']

    return None

# List<...>
def listType(type):
    match = re.match(config.regex['list'], type)
    if match:
        return match.groupdict()['type']

    return None

# [...]
def arrayType(type):
    match = re.match(config.regex['array'], type)
    if match:
        return match.groupdict()['type']

    return None

def constValue(value):
    return re.findall(config.regex['const'], str(value))

def dsl(value):
    matched = re.match(config.regex['dsl'], value)
    if not matched:
        return None

    groups = matched.groupdict()
    header, parameters = groups['header'], groups['parameters']
    parameters = parameters.replace(' ', '').replace('\xa0', '').split(',')

    values = []
    for i in range(len(parameters)):
        parameter = parameters[i]
        argument = config.dsl[header][i]

        values.append(converter.cast(argument['type'], parameter))

    return header, values