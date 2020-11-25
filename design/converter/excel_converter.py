import openpyxl
import json
import validate
import jinja2
import shutil
import os
import logging
import argparse


def sheet_header(sheet):
    result = []
    for i in range(1, sheet.max_column+1):
        name = sheet.cell(1, i).value
        type = sheet.cell(2, i).value
        dest = sheet.cell(3, i).value
        validate.assert_header(name, type, dest)
        result.append({'name': name, 'type': type, 'dest': dest})

    primary = [x['name'] for x in result if x['name'].startswith('*')]
    if len(primary) > 1:
        primaries = ', '.join(primary)
        raise Exception(
            f'cannot contains 2 or more primary key. ({primaries})')

    return result


def sheet_body(sheet, header):
    result = []
    rows = sheet.max_row-3
    for row in range(4, sheet.max_row+1):
        data_set = {}
        for col in range(1, sheet.max_column+1):
            type = header[col-1]['type']
            value = sheet.cell(row, col).value
            value = validate.assert_data(type, value)

            name = header[col-1]['name']
            if name.startswith('*'):
                name = name[1:]
            data_set[name] = value

        result.append(data_set)

    dest_set = {'server': [{} for x in range(rows)], 'client': [
        {} for x in range(rows)]}
    for i in range(len(header)):
        targets = []
        if header[i]['dest'] == 'common':
            targets.append(dest_set['server'])
            targets.append(dest_set['client'])
        elif header[i]['dest'] == 'server':
            targets.append(dest_set['server'])
        else:
            targets.append(dest_set['client'])

        name = header[i]['name']
        if name.startswith('*'):
            name = name[1:]

        for row in range(rows):
            for target in targets:
                target[row][name] = result[row][name]

    return dest_set


def sheet_class(cname, header):
    loader = jinja2.FileSystemLoader('templates')
    env = jinja2.Environment(loader=loader)
    template = env.get_template('template.txt')

    dest_set = {'server': [x for x in header if x['dest'] == 'server' or x['dest'] == 'common'], 'client': [
        x for x in header if x['dest'] == 'server' or x['dest'] == 'common']}

    properties = {'server': [], 'client': []}
    for dest in dest_set:
        for data in dest_set[dest]:
            type = data['type']
            primary = data['name'].startswith('*')
            name = data['name'].replace('*', '')
            if primary:
                properties[dest].append(f'[Key]')
            properties[dest].append(
                f"public {type} {name.capitalize()} {{ get; set; }}")

    for dest in properties:
        properties[dest] = [f'    {x}' for x in properties[dest]]
        properties[dest] = '\n'.join(properties[dest])

    return {'server': template.render({'name': cname, 'property': properties['server']}),
            'client': template.render({'name': cname, 'property': properties['client']})}


def sheet_table(table_name, class_name, sheet_name, header):
    loader = jinja2.FileSystemLoader('templates')
    env = jinja2.Environment(loader=loader)

    result = {'server': None, 'client': None}
    for dest in result:
        key = [x['type']
               for x in header if x['name'].startswith('*') and (x['dest'] == dest or x['dest'] == 'common')]
        if len(key) == 0:
            template = env.get_template('table_list.txt')
            result[dest] = template.render(
                {'sheetName': sheet_name, 'className': class_name, 'tableName': table_name})
        else:
            key = key[0]

            template = env.get_template('table_dict.txt')
            result[dest] = template.render(
                {'sheetName': sheet_name, 'keyType': key, 'className': class_name, 'tableName': table_name})
    return result


if __name__ == '__main__':
    try:
        parser = argparse.ArgumentParser(description='Excel table converter')
        parser.add_argument('--dir', default='..')
        parser.add_argument('--out', default='output')
        args = parser.parse_args()

        if not os.path.isdir(args.dir):
            raise Exception(f'cannot find directory : {args.dir}')

        if os.path.isdir(args.out):
            shutil.rmtree(args.out)

        os.makedirs(f'{args.out}/json/server', exist_ok=True)
        os.makedirs(f'{args.out}/json/client', exist_ok=True)
        os.makedirs(f'{args.out}/class/server', exist_ok=True)
        os.makedirs(f'{args.out}/class/client', exist_ok=True)
        os.makedirs(f'{args.out}/table/server', exist_ok=True)
        os.makedirs(f'{args.out}/table/client', exist_ok=True)
        table_codes = {'server': [], 'client': []}
        for excel in [os.path.join(args.dir, f) for f in os.listdir(args.dir) if os.path.isfile(os.path.join(args.dir, f)) and f.endswith('.xlsx') and not f.startswith('~$')]:
            workbook = openpyxl.load_workbook(excel, data_only=True)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                header = sheet_header(sheet)
                body = sheet_body(sheet, header)

                for dest in body:
                    with open(f'{args.out}/json/{dest}/{sheet_name}.json', 'w', encoding='utf8') as f:
                        f.write(json.dumps(
                            body[dest], indent=4, ensure_ascii=False))

                class_name = sheet_name.capitalize()
                classes = sheet_class(class_name, header)
                for dest in classes:
                    with open(f'{args.out}/class/{dest}/{class_name}.cs', 'w', encoding='utf8') as f:
                        f.write(classes[dest])

                table_name = f'Table{class_name}'
                tcode_pair = sheet_table(
                    table_name, class_name, sheet_name, header)
                table_codes['server'].append(tcode_pair['server'])
                table_codes['client'].append(tcode_pair['client'])

        loader = jinja2.FileSystemLoader('templates')
        env = jinja2.Environment(loader=loader)
        template = env.get_template('table.txt')

        for dest in table_codes:
            table_codes[dest] = '\n'.join(
                [f'    {x}' for x in '\n\n'.join(table_codes[dest]).split('\n')])
            with open(f'{args.out}/table/{dest}/Table.cs', 'w', encoding='utf8') as f:
                f.write(template.render({'contents': table_codes[dest]}))

    except Exception as e:
        logging.error(str(e))
